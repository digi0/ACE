"""Export every ACE dataset to one Excel workbook for human review.

The datasets are committed as JSON because that is what the services read. JSON
is not what a person reviews 1,300 rows in. This writes one .xlsx with a tab per
dataset — filterable, sortable, frozen headers — plus a summary tab saying what
to check and where each row came from.

Read-only: it never touches the JSON.

    python -m backend.data.export_review
    python -m backend.data.export_review --out ~/Desktop/ACE-datasets.xlsx
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent
DEFAULT_OUT = Path.home() / "Desktop" / "ACE-datasets-for-review.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3B2C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)


def _join(value):
    """Lists become readable cells; everything else becomes a string."""
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    return "" if value is None else str(value)


def load(name: str) -> dict:
    path = DATA_DIR / name
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return {}


# (sheet, json file, list key, [(column header, record key)])
SHEETS = [
    ("Clubs", "clubs.json", "clubs", [
        ("Name", "name"), ("Short name", "short_name"),
        ("Categories", "categories"), ("What it is", "summary"),
        ("Instagram", "instagram"), ("Website", "website"),
        ("LinkedIn", "linkedin"), ("Profile page", "url"),
    ]),
    ("Places", "places.json", "places", [
        ("Name", "name"), ("Category", "category"), ("Where", "where"),
        ("What it is", "what_it_is"), ("Good for", "good_for"),
        ("Notes", "notes"), ("Phone", "phone"),
        ("Info page", "url"), ("Directions", "map_url"),
        ("Live hours page", "hours_url"),
    ]),
    ("Events", "events.json", "events", [
        ("Event", "name"), ("Organisation", "organization"),
        ("Starts", "starts_on"), ("Ends", "ends_on"),
        ("Location", "location"), ("Categories", "categories"),
        ("Theme", "theme"), ("Description", "description"),
        ("Event page", "url"), ("Map", "map_url"),
    ]),
    ("Procedures", "procedures.json", "procedures", [
        ("Title", "title"), ("Topic", "topic"),
        ("What it is", "what_it_is"), ("When it applies", "when_to_use"),
        ("Steps", "steps"), ("Forms", "forms"), ("Timing", "timing"),
        ("Who handles it", "who_to_contact"), ("Consequences", "consequences"),
        ("PSU policy", "policy_refs"), ("Source", "source_url"),
    ]),
    ("Money", "money.json", "money", [
        ("Title", "title"), ("Topic", "topic"),
        ("What it is", "what_it_is"), ("Steps", "steps"),
        ("Stated amounts", "amounts"), ("Timing", "timing"),
        ("Who to contact", "who_to_contact"), ("Notes", "notes"),
        ("Source", "source_url"),
    ]),
]

# Columns that hold long prose and need room to breathe.
WIDE = {"What it is", "Steps", "When it applies", "Description", "Notes",
        "Consequences", "Stated amounts"}


def write_sheet(wb, sheet_name, records, columns):
    ws = wb.create_sheet(sheet_name)
    ws.append([label for label, _ in columns])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for record in records:
        ws.append([_join(record.get(key)) for _, key in columns])

    for index, (label, _) in enumerate(columns, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = 60 if label in WIDE else 26
        for cell in ws[letter][1:]:
            cell.alignment = Alignment(wrap_text=label in WIDE, vertical="top")

    ws.freeze_panes = "A2"                       # headers stay put while scrolling
    ws.auto_filter.ref = ws.dimensions           # sort and filter from row 1
    ws.row_dimensions[1].height = 22
    return len(records)


def write_summary(wb, rows):
    ws = wb.create_sheet("Start here", 0)
    ws["A1"] = "ACE datasets — for review"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "One tab per dataset. Every tab has filters on row 1 — click the arrow in "
        "a header to sort or narrow. Nothing here is live yet; none of it is pushed."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 34

    headers = ["Tab", "Rows", "Where it came from", "Worth checking"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        ws.append(row)

    for index, width in enumerate([16, 10, 44, 78], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    return ws


NOTES = {
    "Clubs": (
        "discover.psu.edu (Anthology Engage API)",
        "Are the Instagram/website links right? Contact emails were deliberately "
        "NOT scraped — they are individual students' addresses; the profile page "
        "carries them instead. Say if you want that changed.",
    ),
    "Places": (
        "liveon.psu.edu, libraries.psu.edu, studentaffairs.psu.edu, "
        "transportation.psu.edu, www.it.psu.edu",
        "No opening hours are stored anywhere on purpose — they change by term and "
        "by day, so each row links its live hours page. 'Directions' is a Google "
        "Maps SEARCH, not a verified pin. Check the search actually lands on the "
        "right building.",
    ),
    "Events": (
        "discover.psu.edu (same Engage API as Clubs)",
        "This is the only dataset that expires. ACE stops naming events once the "
        "snapshot is over 3 weeks old. Re-run the scraper weekly.",
    ),
    "Procedures": (
        "registrar.psu.edu + studentpetitions.psu.edu",
        "Highest-stakes tab: these send a student to file real paperwork. Check the "
        "steps and the office against the source link. Written by an LLM reading "
        "the page, so wording may drift from the original.",
    ),
    "Money": (
        "bursar.psu.edu",
        "Navigation only — how billing works and who to contact. No aid advice, no "
        "balances. Check that 'Stated amounts' quotes the page exactly.",
    ),
}


def build(out_path: Path) -> tuple[Path, list]:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    summary_rows = []
    for sheet_name, filename, list_key, columns in SHEETS:
        data = load(filename)
        records = data.get(list_key, [])
        if not records:
            continue
        count = write_sheet(wb, sheet_name, records, columns)
        source, check = NOTES.get(sheet_name, ("", ""))
        scraped = (data.get("scraped_at") or "")[:10]
        summary_rows.append([
            sheet_name, count,
            f"{source}\n(scraped {scraped})" if scraped else source,
            check,
        ])

    write_summary(wb, summary_rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, summary_rows


def main() -> int:
    ap = argparse.ArgumentParser(description="Export ACE datasets to Excel for review.")
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    path, rows = build(Path(args.out).expanduser())
    total = sum(r[1] for r in rows)
    print(f"Wrote {len(rows)} sheet(s), {total} rows → {path}")
    for name, count, *_ in rows:
        print(f"  {name:<12} {count:>5}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
